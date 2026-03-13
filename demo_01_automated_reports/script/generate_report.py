"""
Automated Sales Report Generator
==================================
Reads raw sales data from CSV, validates and cleans it,
calculates key business metrics, and exports a formatted
Excel report, a plain-text summary, and two PNG charts.
"""

import os
import sys
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

INPUT_FILE = INPUT_DIR / "sales_data.csv"
EXCEL_FILE = OUTPUT_DIR / "sales_summary.xlsx"
TEXT_FILE = OUTPUT_DIR / "summary.txt"
CHART_REGION = OUTPUT_DIR / "revenue_by_region.png"
CHART_PRODUCTS = OUTPUT_DIR / "top_products.png"


# ---------------------------------------------------------------------------
# 1. Load data
# ---------------------------------------------------------------------------
def load_data(filepath: Path) -> pd.DataFrame:
    """Load the CSV file and return a raw DataFrame."""
    if not filepath.exists():
        print(f"[ERROR] Input file not found: {filepath}")
        sys.exit(1)

    df = pd.read_csv(filepath, parse_dates=["date"])
    print(f"[INFO] Loaded {len(df)} rows from '{filepath.name}'.")
    return df


# ---------------------------------------------------------------------------
# 2. Clean & validate data
# ---------------------------------------------------------------------------
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Validate types, drop incomplete rows, and recalculate total_value."""
    initial_count = len(df)

    # Drop rows with any missing value in critical columns
    critical_cols = ["date", "client_name", "product", "category",
                     "region", "quantity", "unit_price"]
    df = df.dropna(subset=critical_cols)

    # Enforce correct types
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").astype("Int64")
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce")
    df = df.dropna(subset=["quantity", "unit_price"])

    # Recalculate total_value to guarantee consistency
    df["total_value"] = df["quantity"].astype(float) * df["unit_price"]

    dropped = initial_count - len(df)
    if dropped:
        print(f"[WARN] {dropped} row(s) removed during cleaning.")
    else:
        print("[INFO] Data validation passed — no rows removed.")

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 3. Calculate metrics
# ---------------------------------------------------------------------------
def calculate_metrics(df: pd.DataFrame) -> dict:
    """Compute all required business metrics and return them in a dict."""
    metrics = {}

    metrics["total_revenue"] = df["total_value"].sum()
    metrics["total_units_sold"] = int(df["quantity"].sum())
    metrics["average_order_value"] = df["total_value"].mean()

    metrics["revenue_by_region"] = (
        df.groupby("region")["total_value"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"total_value": "revenue"})
    )

    metrics["revenue_by_category"] = (
        df.groupby("category")["total_value"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"total_value": "revenue"})
    )

    metrics["top_products"] = (
        df.groupby("product")["total_value"]
        .sum()
        .sort_values(ascending=False)
        .head(5)
        .reset_index()
        .rename(columns={"total_value": "revenue"})
    )

    metrics["top_clients"] = (
        df.groupby("client_name")["total_value"]
        .sum()
        .sort_values(ascending=False)
        .head(5)
        .reset_index()
        .rename(columns={"total_value": "revenue"})
    )

    print("[INFO] Metrics calculated successfully.")
    return metrics


# ---------------------------------------------------------------------------
# 4. Export Excel report
# ---------------------------------------------------------------------------
def _style_header_row(ws, header_fill_hex: str = "1F4E79"):
    """Apply bold white text on dark-blue background to the first row."""
    fill = PatternFill(fill_type="solid", fgColor=header_fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_column_width(ws, extra: int = 4):
    """Set each column width to the max content length + padding."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + extra


def _format_currency_column(ws, col_index: int, start_row: int = 2):
    """Apply accounting number format to a column (1-indexed)."""
    col_letter = get_column_letter(col_index)
    for row in ws.iter_rows(min_row=start_row, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.number_format = '#,##0.00'


def export_excel_report(df: pd.DataFrame, metrics: dict, filepath: Path):
    """Write all sheets to the Excel workbook with clean formatting."""
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Sheet 1 — Raw Data
        df_export = df.copy()
        df_export["date"] = df_export["date"].dt.strftime("%Y-%m-%d")
        df_export.to_excel(writer, sheet_name="Raw_Data", index=False)

        # Sheet 2 — Summary
        summary_data = {
            "Metric": [
                "Total Revenue",
                "Total Units Sold",
                "Average Order Value",
                "Best Region",
                "Best Category",
                "Best-Selling Product",
                "Top Client",
            ],
            "Value": [
                f"${metrics['total_revenue']:,.2f}",
                f"{metrics['total_units_sold']:,}",
                f"${metrics['average_order_value']:,.2f}",
                metrics["revenue_by_region"].iloc[0]["region"],
                metrics["revenue_by_category"].iloc[0]["category"],
                metrics["top_products"].iloc[0]["product"],
                metrics["top_clients"].iloc[0]["client_name"],
            ],
        }
        pd.DataFrame(summary_data).to_excel(
            writer, sheet_name="Summary", index=False
        )

        # Sheet 3 — Revenue by Region
        metrics["revenue_by_region"].to_excel(
            writer, sheet_name="Revenue_By_Region", index=False
        )

        # Sheet 4 — Revenue by Category
        metrics["revenue_by_category"].to_excel(
            writer, sheet_name="Revenue_By_Category", index=False
        )

        # Sheet 5 — Top Products
        metrics["top_products"].to_excel(
            writer, sheet_name="Top_Products", index=False
        )

        # Sheet 6 — Top Clients
        metrics["top_clients"].to_excel(
            writer, sheet_name="Top_Clients", index=False
        )

    # Post-processing: apply styles via openpyxl
    wb = load_workbook(filepath)

    currency_sheets = {
        "Revenue_By_Region": [2],
        "Revenue_By_Category": [2],
        "Top_Products": [2],
        "Top_Clients": [2],
        "Raw_Data": [7, 8],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _style_header_row(ws)
        _auto_column_width(ws)
        ws.row_dimensions[1].height = 20

        if sheet_name in currency_sheets:
            for col_idx in currency_sheets[sheet_name]:
                _format_currency_column(ws, col_idx)

    wb.save(filepath)
    print(f"[INFO] Excel report saved → {filepath.name}")


# ---------------------------------------------------------------------------
# 5. Export text summary
# ---------------------------------------------------------------------------
def export_text_summary(metrics: dict, filepath: Path):
    """Write a short business-style summary to a plain-text file."""
    best_region = metrics["revenue_by_region"].iloc[0]
    best_category = metrics["revenue_by_category"].iloc[0]
    best_product = metrics["top_products"].iloc[0]
    best_client = metrics["top_clients"].iloc[0]

    lines = [
        "=" * 56,
        "         AUTOMATED SALES REPORT — SUMMARY",
        "=" * 56,
        "",
        f"  Total Revenue        : ${metrics['total_revenue']:>12,.2f}",
        f"  Total Units Sold     : {metrics['total_units_sold']:>12,}",
        f"  Average Order Value  : ${metrics['average_order_value']:>12,.2f}",
        "",
        "  --- Regional Performance ---",
        f"  Best Region          : {best_region['region']}"
        f"  (${best_region['revenue']:,.2f})",
        "",
        "  --- Category Performance ---",
        f"  Best Category        : {best_category['category']}"
        f"  (${best_category['revenue']:,.2f})",
        "",
        "  --- Product Performance ---",
        f"  Best-Selling Product : {best_product['product']}"
        f"  (${best_product['revenue']:,.2f})",
        "",
        "  --- Client Performance ---",
        f"  Top Client           : {best_client['client_name']}"
        f"  (${best_client['revenue']:,.2f})",
        "",
        "=" * 56,
    ]

    filepath.write_text("\n".join(lines), encoding="utf-8")
    print(f"[INFO] Text summary saved  → {filepath.name}")


# ---------------------------------------------------------------------------
# 6. Generate charts
# ---------------------------------------------------------------------------
def _apply_chart_style(ax, title: str, xlabel: str, ylabel: str):
    """Apply a clean, professional style to a matplotlib Axes object."""
    ax.set_title(title, fontsize=14, fontweight="bold", pad=14)
    ax.set_xlabel(xlabel, fontsize=11)
    ax.set_ylabel(ylabel, fontsize=11)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: f"${x:,.0f}")
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", labelsize=10)
    ax.grid(axis="y", linestyle="--", alpha=0.5)


def generate_charts(metrics: dict, region_path: Path, products_path: Path):
    """Produce and save the two PNG charts."""
    palette = ["#1F4E79", "#2E75B6", "#4472C4", "#70AD47", "#ED7D31"]

    # Chart 1 — Revenue by Region (horizontal bar)
    region_df = metrics["revenue_by_region"].sort_values("revenue")
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.barh(
        region_df["region"],
        region_df["revenue"],
        color=palette[:len(region_df)],
        edgecolor="white",
    )
    ax.bar_label(bars, fmt="$%.0f", padding=6, fontsize=9)
    _apply_chart_style(ax, "Revenue by Region", "Revenue (USD)", "Region")
    ax.xaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: f"${x:,.0f}")
    )
    plt.tight_layout()
    fig.savefig(region_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"[INFO] Chart saved         → {region_path.name}")

    # Chart 2 — Top 5 Products by Revenue (vertical bar)
    prod_df = metrics["top_products"].sort_values("revenue", ascending=False)
    fig, ax = plt.subplots(figsize=(9, 5))
    bars = ax.bar(
        prod_df["product"],
        prod_df["revenue"],
        color=palette,
        edgecolor="white",
        width=0.6,
    )
    ax.bar_label(bars, fmt="$%.0f", padding=4, fontsize=9)
    _apply_chart_style(ax, "Top 5 Products by Revenue", "Product", "Revenue (USD)")
    plt.xticks(rotation=20, ha="right")
    plt.tight_layout()
    fig.savefig(products_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"[INFO] Chart saved         → {products_path.name}")


# ---------------------------------------------------------------------------
# 7. Main
# ---------------------------------------------------------------------------
def main():
    print("\n" + "=" * 56)
    print("  Automated Sales Report Generator — Starting")
    print("=" * 56 + "\n")

    df_raw = load_data(INPUT_FILE)
    df_clean = clean_data(df_raw)
    metrics = calculate_metrics(df_clean)

    export_excel_report(df_clean, metrics, EXCEL_FILE)
    export_text_summary(metrics, TEXT_FILE)
    generate_charts(metrics, CHART_REGION, CHART_PRODUCTS)

    print("\n" + "=" * 56)
    print("  All outputs saved to: output/")
    print("=" * 56 + "\n")


if __name__ == "__main__":
    main()
